"""
CIPP/E Legal SaaS — FastAPI Backend
Serves legal documents from GCS with signed URLs, text extraction,
and highlight position mapping for the Legal Research Lab.
"""

import os
import re
import time
import json
import asyncio
import logging
import datetime
import unicodedata
from contextlib import asynccontextmanager
from typing import Optional, List

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, Request, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, HTMLResponse, FileResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy.orm import Session

from gcs_client import GCSLegalClient
from database import (
    init_db, get_db,
    create_upload, list_uploads, update_status, delete_upload,
    save_embeddings, load_embeddings, is_indexed,
)
from embedding_service import init_embedding, index_document, semantic_search, embed_query
import docai_service

# ── Vertex AI (google-genai) — optional, graceful fallback if not installed ──
try:
    from google import genai as _genai
    from google.genai import types as _genai_types
    _VERTEX_AI_AVAILABLE = True
except ImportError:
    _VERTEX_AI_AVAILABLE = False
    logging.getLogger("cipp-legal-api").warning("google-genai not installed — /api/documents/vertex-analyze will return 503")

# ── Config ──────────────────────────────────────────────────────────
load_dotenv()

# Supporta credenziali GCS sia da file locale che da variabile d'ambiente JSON
# (necessario per deployment su Render/cloud dove non si può caricare file)
_CREDS_JSON_ENV = os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON", "")
if _CREDS_JSON_ENV:
    _tmp_creds = "/tmp/gcs_service_account.json"
    with open(_tmp_creds, "w") as _f:
        _f.write(_CREDS_JSON_ENV)
    CREDENTIALS_PATH = _tmp_creds
else:
    CREDENTIALS_PATH = os.path.join(
        os.path.dirname(__file__),
        os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "credentials/service_account.json")
    )

BUCKET_NAME        = os.getenv("GCS_BUCKET_NAME", "data-protection-archive")
CORS_ORIGINS       = os.getenv("CORS_ORIGINS", "http://localhost:7890").split(",")
VERTEX_PROJECT     = os.getenv("GOOGLE_CLOUD_PROJECT", "")
VERTEX_LOCATION    = os.getenv("VERTEX_AI_LOCATION", "us-central1")
VERTEX_MODEL       = os.getenv("VERTEX_AI_MODEL", "gemini-2.0-flash")

# Ordered fallback chain for the Gemini proxy — first model that responds wins
_PROXY_FALLBACK_MODELS = [
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-1.5-flash",
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
    "gemini-1.5-pro",
]
GEMINI_API_KEY     = os.getenv("GEMINI_API_KEY", "")  # Non-Vertex fallback

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("cipp-legal-api")

# ── App lifecycle ───────────────────────────────────────────────────
gcs = None  # type: Optional[GCSLegalClient]


@asynccontextmanager
async def lifespan(app: FastAPI):
    global gcs
    logger.info("🔐 Inizializzazione GCS client...")
    logger.info(f"   Bucket: {BUCKET_NAME}")
    logger.info(f"   Credentials: {CREDENTIALS_PATH}")
    try:
        gcs = GCSLegalClient(CREDENTIALS_PATH, BUCKET_NAME)
        logger.info("✅ GCS client inizializzato con successo")
    except Exception as e:
        logger.error(f"❌ Errore inizializzazione GCS: {e}")
        gcs = None
    # ── DB init ──────────────────────────────────────────────────────
    try:
        init_db()
        logger.info("✅ Database uploads inizializzato")
    except Exception as e:
        logger.error(f"❌ Errore init DB: {e}")
    # ── Embedding service init ────────────────────────────────────────
    try:
        init_embedding(GEMINI_API_KEY)
    except Exception as e:
        logger.error(f"❌ Errore init embedding service: {e}")
    # ── Document AI OCR init ─────────────────────────────────────────
    try:
        docai_service.init_docai(CREDENTIALS_PATH)
    except Exception as e:
        logger.error(f"❌ Errore init Document AI: {e}")
    yield
    logger.info("🛑 Shutdown server")


app = FastAPI(
    title="CIPP/E Legal SaaS API",
    description="Backend API per il Legal Research Lab — gestione documenti GCS con signed URLs",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Pydantic Models ────────────────────────────────────────────────

class HighlightRequest(BaseModel):
    blob_name: str
    search_texts: List[str]
    page_hint: Optional[int] = None   # suggested page to search first
    strict_page: bool = False          # if True, only search page_hint — no global fallback


class TextExtractionRequest(BaseModel):
    blob_name: str
    page_start: int = 0
    page_end: Optional[int] = None


class VertexAnalyzeRequest(BaseModel):
    blob_name: str
    question: str = "Analizza questo documento per la certificazione CIPP/E IAPP"
    model: Optional[str] = None  # overrides VERTEX_MODEL env var if set


class UploadRequestBody(BaseModel):
    user_id: str
    filename: str
    content_type: str = "application/pdf"
    file_size: Optional[int] = None         # bytes, per la validazione
    expiration_minutes: int = 10


class UploadConfirmBody(BaseModel):
    user_id: str
    blob_name: str
    original_filename: str
    file_size: Optional[int] = None
    mime_type: Optional[str] = None


class StatusUpdateBody(BaseModel):
    status: str                              # in_attesa | approvato | rifiutato
    notes: Optional[str] = None


# ── Utility ─────────────────────────────────────────────────────────

def _sanitize_filename(name: str) -> str:
    """Normalizza un filename: rimuove caratteri non-ASCII, spazi → underscore."""
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    name = re.sub(r"[^\w.\-]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name or "file"


# ── Health ──────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    return {
        "status": "ok",
        "gcs_connected": gcs is not None,
        "bucket": BUCKET_NAME
    }


# ── PDF Listing ─────────────────────────────────────────────────────

@app.get("/api/documents")
async def list_documents(
    prefix: str = Query("", description="Filtra per prefisso cartella"),
    limit: int = Query(200, description="Numero massimo di risultati")
):
    """Lista tutti i PDF nel bucket GCS con metadati."""
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        docs = gcs.list_pdfs(prefix=prefix, max_results=limit)
        return {
            "count": len(docs),
            "bucket": BUCKET_NAME,
            "documents": docs
        }
    except Exception as e:
        logger.error(f"Errore listing documenti: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── PDF Stream proxy (CORS-safe for browser/PDF.js) ─────────────────

_PDF_CACHE: dict = {}          # blob_name → {"data": bytes, "ts": float}
_PDF_CACHE_TTL = 600           # 10 min — riuso tra richieste successive nella stessa sessione

def _pdf_bytes_cached(blob_name: str) -> bytes:
    """Scarica il PDF da GCS con cache in memoria (TTL 10 min)."""
    now = time.time()
    entry = _PDF_CACHE.get(blob_name)
    if entry and now - entry["ts"] < _PDF_CACHE_TTL:
        return entry["data"]
    data = gcs.download_pdf_bytes(blob_name)
    _PDF_CACHE[blob_name] = {"data": data, "ts": now}
    return data


@app.get("/api/documents/pdf")
async def stream_pdf(
    request: Request,
    blob_name: str = Query(..., description="Nome del blob nel bucket")
):
    """Serve il PDF con range request support — PDF.js può renderizzare p.1 senza scaricare tutto."""
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        pdf_bytes = _pdf_bytes_cached(blob_name)
        total = len(pdf_bytes)
        filename = blob_name.split("/")[-1]

        range_header = request.headers.get("range", "")
        if range_header:
            m = re.match(r"bytes=(\d+)-(\d*)", range_header)
            if m:
                start = int(m.group(1))
                end = int(m.group(2)) if m.group(2) else total - 1
                end = min(end, total - 1)
                chunk = pdf_bytes[start: end + 1]
                return Response(
                    content=chunk,
                    status_code=206,
                    media_type="application/pdf",
                    headers={
                        "Content-Range": f"bytes {start}-{end}/{total}",
                        "Content-Length": str(len(chunk)),
                        "Accept-Ranges": "bytes",
                        "Cache-Control": "private, max-age=600",
                    },
                )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'inline; filename="{filename}"',
                "Content-Length": str(total),
                "Accept-Ranges": "bytes",
                "Cache-Control": "private, max-age=600",
            },
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore streaming PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Signed URL ──────────────────────────────────────────────────────

@app.get("/api/documents/signed-url")
async def get_signed_url(
    blob_name: str = Query(..., description="Nome del blob nel bucket"),
    expiration: int = Query(60, description="Durata URL in minuti", ge=5, le=720)
):
    """Genera un Signed URL per accesso privato al PDF."""
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        url = gcs.generate_signed_url(blob_name, expiration_minutes=expiration)
        return {
            "blob_name": blob_name,
            "signed_url": url,
            "expires_in_minutes": expiration
        }
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore generazione signed URL: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Text Extraction ────────────────────────────────────────────────

@app.post("/api/documents/extract-text")
async def extract_text(req: TextExtractionRequest):
    """Estrai testo dal PDF con granularità per pagina."""
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        result = gcs.extract_text(
            req.blob_name,
            page_start=req.page_start,
            page_end=req.page_end
        )
        return result
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore estrazione testo: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Highlight Positions ────────────────────────────────────────────

@app.post("/api/documents/highlights")
async def find_highlights(req: HighlightRequest):
    """
    Trova le posizioni esatte (coordinate rettangolari) delle frasi
    nel PDF per l'evidenziazione lato frontend.
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        all_matches = []
        for text in req.search_texts:
            result = gcs.extract_text_with_positions(
                req.blob_name, text,
                page_hint=req.page_hint,
                strict_page=req.strict_page
            )
            all_matches.extend(result["matches"])
        return {
            "blob_name": req.blob_name,
            "total_matches": len(all_matches),
            "highlights": all_matches
        }
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore ricerca highlights: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── PDF Metadata ────────────────────────────────────────────────────

@app.get("/api/documents/metadata")
async def get_metadata(
    blob_name: str = Query(..., description="Nome del blob nel bucket")
):
    """Restituisce i metadati del PDF (titolo, autore, pagine, ecc.)."""
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        meta = gcs.get_pdf_metadata(blob_name)
        return meta
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore lettura metadati: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Document Context for AI ────────────────────────────────────────

@app.get("/api/documents/ai-context")
async def get_ai_context(
    blob_name: str = Query(..., description="Nome del blob nel bucket"),
    max_pages: int = Query(30, description="Numero massimo di pagine da estrarre")
):
    """
    Estrae il testo del documento e lo prepara come contesto
    per l'analisi AI (Legal Research Lab).
    Restituisce il testo strutturato + metadati per il system prompt.
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        meta = gcs.get_pdf_metadata(blob_name)
        text_data = gcs.extract_text(blob_name, page_start=0, page_end=max_pages)

        return {
            "metadata": meta,
            "text": text_data["full_text"],
            "pages_extracted": len(text_data["pages"]),
            "total_pages": text_data["total_pages"],
            "total_chars": text_data["total_chars"],
            "truncated": max_pages < text_data["total_pages"]
        }
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore preparazione contesto AI: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── PDF Proxy (avoids CORS issues with GCS signed URLs) ────────────

@app.get("/api/documents/pdf-proxy")
async def pdf_proxy(
    blob_name: str = Query(..., description="Nome del blob nel bucket")
):
    """
    Proxy the PDF binary through the backend to avoid CORS issues.
    PDF.js loads from same-origin /api/documents/pdf-proxy?blob_name=...
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    try:
        pdf_bytes = gcs.download_pdf_bytes(blob_name)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename=\"{blob_name}\"",
                "Cache-Control": "public, max-age=3600"
            }
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{blob_name}' non trovato")
    except Exception as e:
        logger.error(f"Errore proxy PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Vertex AI Deep Analysis ─────────────────────────────────────────

@app.post("/api/documents/vertex-analyze")
async def vertex_analyze(req: VertexAnalyzeRequest):
    """
    Analizza un PDF direttamente via Vertex AI (Gemini).
    Il modello legge il PDF raw da GCS e restituisce analisi strutturata
    con numero di pagina esatto e coordinata y_start (scala 0-1000) per
    ciascuna evidenza legale — elimina la necessità di text-search fuzzy.

    Risposta JSON:
    {
      "analisi": "...",
      "evidenze_documentali": [
        { "num": 1, "concetto": "...", "pagina": N, "y_start": 0-1000,
          "testo_esatto": "...", "spiegazione": "..." },
        ...
      ]
    }
    """
    if not _VERTEX_AI_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="google-genai non installato — esegui: pip install google-genai"
        )
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    model_id = req.model or VERTEX_MODEL

    # Verify blob exists before calling Vertex AI (cheap GCS call)
    try:
        blob = gcs._bucket.blob(req.blob_name)
        if not blob.exists():
            raise FileNotFoundError(req.blob_name)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"GCS check error: {e}")

    prompt = (
        "Sei un esperto Senior di diritto europeo della protezione dei dati (GDPR, ePrivacy, AI Act).\n\n"
        f"Analizza questo documento PDF e rispondi alla seguente richiesta:\n{req.question}\n\n"
        "FORMATO DI RISPOSTA OBBLIGATORIO — Rispondi SOLO con JSON valido (no markdown, no testo extra):\n"
        "{\n"
        '  "analisi": "Sintesi generale del documento in 3-4 frasi",\n'
        '  "evidenze_documentali": [\n'
        "    {\n"
        '      "num": 1,\n'
        '      "concetto": "Titolo breve del concetto chiave (max 6 parole)",\n'
        '      "pagina": <numero_pagina_intero_1_indexed>,\n'
        '      "y_start": <posizione_verticale_0_a_1000>,\n'
        '      "testo_esatto": "Frase letterale copiata dal documento (max 120 caratteri)",\n'
        '      "spiegazione": "Spiegazione con riferimento normativo es. [Art. 5(1)(a) GDPR] (max 200 caratteri)"\n'
        "    }\n"
        "  ]\n"
        "}\n\n"
        "REGOLE:\n"
        "- evidenze_documentali deve contenere esattamente 6-8 elementi\n"
        "- pagina = numero di pagina (1-indexed) dove si trova fisicamente il testo\n"
        "- y_start = posizione verticale sulla pagina, scala 0-1000 (0=cima, 1000=fondo)\n"
        "- testo_esatto deve essere copiato LETTERALMENTE dal documento, senza modifiche\n"
        "- spiegazione include riferimento normativo preciso tra parentesi quadre\n"
        "- Rispondi SOLO con il JSON, assolutamente nessun testo prima o dopo"
    )

    try:
        # ── Strategy 1: Gemini API key (no Vertex AI IAM needed) ──────────
        # Downloads PDF bytes from GCS, sends as inline data to Gemini REST API.
        # Preferred when GEMINI_API_KEY is set.
        if GEMINI_API_KEY:
            pdf_bytes = gcs.download_pdf_bytes(req.blob_name)
            pdf_part  = _genai_types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf")
            client    = _genai.Client(api_key=GEMINI_API_KEY)
            logger.info(f"[VertexAI] Using Gemini API key (inline PDF, model={model_id})")
        else:
            # ── Strategy 2: Vertex AI via gs:// URI (needs aiplatform.user role) ──
            pdf_uri  = f"gs://{BUCKET_NAME}/{req.blob_name}"
            pdf_part = _genai_types.Part.from_uri(file_uri=pdf_uri, mime_type="application/pdf")
            client   = _genai.Client(vertexai=True, project=VERTEX_PROJECT, location=VERTEX_LOCATION)
            logger.info(f"[VertexAI] Using Vertex AI (gs:// URI, model={model_id})")

        text_part = _genai_types.Part.from_text(text=prompt)

        _va_contents = [_genai_types.Content(role="user", parts=[pdf_part, text_part])]
        _va_config = _genai_types.GenerateContentConfig(
            response_mime_type="application/json",
            max_output_tokens=4096,
            temperature=0.1,
        )
        def _va_call():
            return client.models.generate_content(
                model=model_id, contents=_va_contents, config=_va_config)
        response = await asyncio.to_thread(_va_call)

        raw = response.text.strip()
        # Strip markdown code fences if the model wrapped the JSON anyway
        if raw.startswith("```"):
            raw = raw.split("```", 2)[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        result = json.loads(raw)

        # Validate & sanitise
        if "evidenze_documentali" not in result:
            result["evidenze_documentali"] = []
        for i, ev in enumerate(result["evidenze_documentali"]):
            ev.setdefault("num",        i + 1)
            ev.setdefault("concetto",   f"Evidenza {i + 1}")
            ev.setdefault("pagina",     1)
            ev.setdefault("y_start",    500)
            ev.setdefault("testo_esatto", "")
            ev.setdefault("spiegazione",  "")
            # Clamp y_start to [0, 1000]
            ev["y_start"] = max(0, min(1000, int(ev["y_start"])))

        logger.info(
            f"[VertexAI] {req.blob_name} → {len(result['evidenze_documentali'])} evidenze, model={model_id}"
        )
        return result

    except json.JSONDecodeError as e:
        logger.error(f"[VertexAI] JSON parse error: {e} — raw: {raw[:300]}")
        raise HTTPException(status_code=502, detail=f"Vertex AI JSON parse error: {e}")
    except Exception as e:
        logger.error(f"[VertexAI] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Semantic Embedding ─────────────────────────────────────────────────────────

class EmbedRequest(BaseModel):
    blob_name: str
    force: bool = False   # se True, re-indicizza anche se già presente


@app.post("/api/documents/embed")
async def embed_document(req: EmbedRequest, db: Session = Depends(get_db)):
    """
    Indicizza un documento GCS generando embedding semantici per tutti i chunk.
    Idempotente: se il documento è già indicizzato ritorna il conteggio esistente
    senza ricalcolare (usa force=true per forzare).
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    if not is_indexed(db, req.blob_name) or req.force:
        try:
            # Scarica PDF bytes
            pdf_bytes = gcs.download_pdf_bytes(req.blob_name)

            # Auto-detect: PDF scansionato → Document AI OCR, altrimenti PyMuPDF
            if docai_service.is_ready() and docai_service.needs_ocr(pdf_bytes):
                logger.info(f"[Embed] PDF scansionato rilevato — uso Document AI OCR per {req.blob_name}")
                raw_blocks = docai_service.ocr_pdf_bytes(pdf_bytes)
                # Converte blocchi OCR nel formato atteso da save_embeddings
                from embedding_service import embed_chunks
                texts = [b["text"] for b in raw_blocks]
                embeddings = embed_chunks(texts)
                chunks = [
                    {
                        "chunk_index": i,
                        "text":        blk["text"],
                        "char_start":  blk.get("char_start", 0),
                        "page_number": blk.get("page_number"),
                        "article_ref": blk.get("article_ref"),
                        "coordinates": blk.get("coordinates"),
                        "embedding":   emb,
                    }
                    for i, (blk, emb) in enumerate(zip(raw_blocks, embeddings))
                ]
            else:
                # Percorso normale: PyMuPDF estrae blocchi con coordinate e numero pagina
                chunks = index_document(pdf_bytes)

            if not chunks:
                raise HTTPException(status_code=422, detail="Nessun blocco testuale estraibile dal PDF")
            count = save_embeddings(db, req.blob_name, chunks)
            pages = max((c.get("page_number") or 0) for c in chunks)
            arts  = len(set(c["article_ref"] for c in chunks if c.get("article_ref")))
            logger.info(f"[Embed] {req.blob_name} → {count} chunk, {pages} pagine, {arts} art. refs")
            return {
                "blob_name":      req.blob_name,
                "chunks_indexed": count,
                "pages_covered":  pages,
                "article_refs":   arts,
                "cached":         False,
            }
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"[Embed] Errore indicizzazione {req.blob_name}: {e}")
            raise HTTPException(status_code=500, detail=str(e))
    else:
        rows  = load_embeddings(db, req.blob_name)
        pages = max((r.get("page_number") or 0) for r in rows) if rows else 0
        arts  = len(set(r["article_ref"] for r in rows if r.get("article_ref")))
        return {
            "blob_name":      req.blob_name,
            "chunks_indexed": len(rows),
            "pages_covered":  pages,
            "article_refs":   arts,
            "cached":         True,
        }


class SemanticSearchRequest(BaseModel):
    blob_name: str
    query: str
    top_k: int = 5
    min_score: float = 0.55
    auto_index: bool = True   # se True, indicizza automaticamente se non presente


@app.post("/api/documents/semantic-search")
async def semantic_search_endpoint(req: SemanticSearchRequest, db: Session = Depends(get_db)):
    """
    Ricerca semantica nei chunk di un documento.
    Ritorna i top_k chunk più rilevanti per la query con score di similarità.
    Se auto_index=True e il documento non è ancora indicizzato, lo indicizza ora.
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")
    # Auto-index se necessario
    if not is_indexed(db, req.blob_name):
        if not req.auto_index:
            raise HTTPException(status_code=404, detail="Documento non ancora indicizzato. Chiama /api/documents/embed prima.")
        try:
            pdf_bytes = gcs.download_pdf_bytes(req.blob_name)
            chunks    = index_document(pdf_bytes)
            save_embeddings(db, req.blob_name, chunks)
            logger.info(f"[Embed] Auto-index {req.blob_name} → {len(chunks)} chunk")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Auto-index fallito: {e}")

    stored  = load_embeddings(db, req.blob_name)
    results = semantic_search(req.query, stored, top_k=req.top_k, min_score=req.min_score)

    # Testo combinato con score visibile all'AI (per RAG coverage assessment)
    combined_text = "\n\n".join(
        f"[Rilevanza: {round(r['score'] * 100)}%] {r['text']}"
        for r in results
    )

    # Highlights pronti per PDF.js: page_number + coordinates + article_ref
    highlights = [
        {
            "text":        r["text"][:120],   # snippet per highlight visivo
            "page":        r["page_number"],
            "article_ref": r["article_ref"],
            "score":       r["score"],
            "coordinates": r.get("coordinates"),
        }
        for r in results if r.get("page_number")
    ]

    return {
        "blob_name":        req.blob_name,
        "query":            req.query,
        "results":          results,
        "highlights":       highlights,
        "combined_text":    combined_text,
        "total_chunks":     len(stored),
        "semantic_available": True,
    }


# ── Document AI OCR + Chunks ───────────────────────────────────────────────────

class OcrRequest(BaseModel):
    blob_name: str
    force: bool = False   # se True, esegue OCR anche su PDF con testo selezionabile


class OcrEmbedRequest(BaseModel):
    blob_name: str
    force_ocr: bool = False    # forza OCR anche se PyMuPDF estrae testo
    force_reindex: bool = False # forza re-indicizzazione anche se già presente


@app.get("/api/documents/ocr-status")
async def ocr_status():
    """Verifica se Document AI OCR è disponibile."""
    return {
        "available":    docai_service.is_ready(),
        "project_id":   docai_service.DOCAI_PROJECT,
        "location":     docai_service.DOCAI_LOCATION,
        "processor_id": docai_service.DOCAI_PROCESSOR,
    }


@app.post("/api/documents/ocr")
async def ocr_document(req: OcrRequest):
    """
    Esegue OCR su un documento GCS tramite Google Cloud Document AI.

    Per default rileva automaticamente se il PDF è scansionato (PyMuPDF < 80 chars/pagina).
    Con force=true esegue OCR comunque, utile per documenti con testo incorporato ma
    qualità scarsa (rendering anomalo, font non standard).

    Risposta:
      {
        "blob_name": "...",
        "blocks": [ {text, page_number, article_ref, coordinates, char_start, source}, ... ],
        "total_blocks": N,
        "pages": N,
        "ocr_used": true
      }
    """
    if not docai_service.is_ready():
        raise HTTPException(
            status_code=503,
            detail="Document AI non inizializzato. Verifica DOCAI_PROJECT_ID/DOCAI_LOCATION/DOCAI_PROCESSOR_ID nel .env e che google-cloud-documentai sia installato."
        )
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    try:
        pdf_bytes = gcs.download_pdf_bytes(req.blob_name)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore download PDF: {e}")

    # Auto-detect: se il PDF ha già testo selezionabile e non è forzato, avverti
    ocr_needed = req.force or docai_service.needs_ocr(pdf_bytes)
    if not ocr_needed:
        return {
            "blob_name":    req.blob_name,
            "blocks":       [],
            "total_blocks": 0,
            "pages":        0,
            "ocr_used":     False,
            "message":      (
                "Il PDF contiene già testo selezionabile — OCR non necessario. "
                "Usa force=true per forzare comunque l'OCR, oppure chiama "
                "/api/documents/embed per l'indicizzazione normale via PyMuPDF."
            ),
        }

    try:
        blocks = docai_service.ocr_pdf_bytes(pdf_bytes)
        pages  = max((b.get("page_number") or 0) for b in blocks) if blocks else 0
        return {
            "blob_name":    req.blob_name,
            "blocks":       blocks,
            "total_blocks": len(blocks),
            "pages":        pages,
            "ocr_used":     True,
        }
    except Exception as e:
        logger.error(f"[DocAI] Errore OCR {req.blob_name}: {e}")
        raise HTTPException(status_code=500, detail=f"Document AI OCR error: {e}")


@app.post("/api/documents/ocr-embed")
async def ocr_embed_document(req: OcrEmbedRequest, db: Session = Depends(get_db)):
    """
    Pipeline completa OCR → Embedding per documenti scansionati.

    Flusso:
      1. Scarica PDF da GCS
      2. Rileva automaticamente se serve OCR (o usa force_ocr=true)
         - PDF scansionato → Document AI OCR → blocchi con coordinate
         - PDF con testo   → PyMuPDF normale (stessa logica di /api/documents/embed)
      3. Genera embedding Gemini per ogni blocco
      4. Salva nel DB (idempotente: salta se già indicizzato, a meno di force_reindex=true)

    Risposta identica a /api/documents/embed:
      { blob_name, chunks_indexed, pages_covered, article_refs, cached, ocr_used }
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    # Controlla se già indicizzato
    if is_indexed(db, req.blob_name) and not req.force_reindex:
        rows  = load_embeddings(db, req.blob_name)
        pages = max((r.get("page_number") or 0) for r in rows) if rows else 0
        arts  = len(set(r["article_ref"] for r in rows if r.get("article_ref")))
        return {
            "blob_name":      req.blob_name,
            "chunks_indexed": len(rows),
            "pages_covered":  pages,
            "article_refs":   arts,
            "cached":         True,
            "ocr_used":       False,
        }

    try:
        pdf_bytes = gcs.download_pdf_bytes(req.blob_name)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore download PDF: {e}")

    # ── Scegli extraction strategy ─────────────────────────────────────────────
    ocr_used = False

    if req.force_ocr or docai_service.needs_ocr(pdf_bytes):
        if not docai_service.is_ready():
            raise HTTPException(
                status_code=503,
                detail=(
                    "Il PDF sembra scansionato ma Document AI non è disponibile. "
                    "Verifica la configurazione DOCAI_* nel .env."
                )
            )
        logger.info(f"[OcrEmbed] Usando Document AI OCR per {req.blob_name}")
        try:
            blocks = docai_service.ocr_pdf_bytes(pdf_bytes)
            ocr_used = True
        except Exception as e:
            logger.error(f"[OcrEmbed] OCR fallito, fallback a PyMuPDF: {e}")
            blocks = index_document(pdf_bytes)   # fallback
    else:
        logger.info(f"[OcrEmbed] Usando PyMuPDF per {req.blob_name}")
        blocks = index_document(pdf_bytes)

    if not blocks:
        raise HTTPException(
            status_code=422,
            detail="Nessun blocco testuale estraibile dal documento (né PyMuPDF né OCR)"
        )

    # ── Genera embedding e salva ────────────────────────────────────────────────
    from embedding_service import embed_chunks

    texts      = [b["text"] for b in blocks]
    embeddings = embed_chunks(texts)

    chunks = []
    for i, (blk, emb) in enumerate(zip(blocks, embeddings)):
        chunks.append({
            "chunk_index": i,
            "text":        blk["text"],
            "char_start":  blk.get("char_start", 0),
            "page_number": blk.get("page_number"),
            "article_ref": blk.get("article_ref"),
            "coordinates": blk.get("coordinates"),
            "embedding":   emb,
        })

    count = save_embeddings(db, req.blob_name, chunks)
    pages = max((c.get("page_number") or 0) for c in chunks)
    arts  = len(set(c["article_ref"] for c in chunks if c.get("article_ref")))

    logger.info(
        f"[OcrEmbed] {req.blob_name} → {count} chunk, {pages} pagine, "
        f"{arts} art.refs, ocr={ocr_used}"
    )
    return {
        "blob_name":      req.blob_name,
        "chunks_indexed": count,
        "pages_covered":  pages,
        "article_refs":   arts,
        "cached":         False,
        "ocr_used":       ocr_used,
    }


# ── Document AI: Ingestione SOLO DocAI + ritorno chunks ────────────────────────

class DocAIIngestRequest(BaseModel):
    blob_name: str
    force: bool = False   # forza re-ingestione anche se già presente


@app.post("/api/documents/docai-ingest")
async def docai_ingest(req: DocAIIngestRequest, db: Session = Depends(get_db)):
    """
    Processa il PDF con Google Cloud Document AI (sempre — ignora PyMuPDF).
    Estrae blocchi con coordinate spaziali normalizzate, genera embedding Gemini
    e salva nel DB. Idempotente con force=false.

    Questo è il punto di ingresso per il sistema pixel-perfect:
    i chunk salvati qui vengono poi usati da /api/documents/docai-chunks
    per il rendering geometrico nel frontend.
    """
    if not docai_service.is_ready():
        raise HTTPException(
            status_code=503,
            detail="Document AI non disponibile. Verifica DOCAI_PROJECT_ID/DOCAI_LOCATION/DOCAI_PROCESSOR_ID nel .env."
        )
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    # Controlla se già indicizzato
    if is_indexed(db, req.blob_name) and not req.force:
        rows  = load_embeddings(db, req.blob_name)
        pages = max((r.get("page_number") or 0) for r in rows) if rows else 0
        arts  = len(set(r["article_ref"] for r in rows if r.get("article_ref")))
        return {
            "blob_name":      req.blob_name,
            "chunks_indexed": len(rows),
            "pages_covered":  pages,
            "article_refs":   arts,
            "cached":         True,
            "ocr_used":       True,
        }

    try:
        pdf_bytes = gcs.download_pdf_bytes(req.blob_name)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Documento '{req.blob_name}' non trovato")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore download PDF: {e}")

    # Document AI OCR — sempre, senza auto-detect
    try:
        blocks = docai_service.ocr_pdf_bytes(pdf_bytes)
    except Exception as e:
        logger.error(f"[DocAI Ingest] OCR fallito per {req.blob_name}: {e}")
        raise HTTPException(status_code=500, detail=f"Document AI error: {e}")

    if not blocks:
        raise HTTPException(status_code=422, detail="Document AI non ha restituito blocchi di testo")

    # Genera embedding e salva
    from embedding_service import embed_chunks
    texts      = [b["text"] for b in blocks]
    embeddings = embed_chunks(texts)

    chunks = [
        {
            "chunk_index": i,
            "text":        blk["text"],
            "char_start":  blk.get("char_start", 0),
            "page_number": blk.get("page_number"),
            "article_ref": blk.get("article_ref"),
            "coordinates": blk.get("coordinates"),
            "embedding":   emb,
        }
        for i, (blk, emb) in enumerate(zip(blocks, embeddings))
    ]

    count = save_embeddings(db, req.blob_name, chunks)
    pages = max((c.get("page_number") or 0) for c in chunks)
    arts  = len(set(c["article_ref"] for c in chunks if c.get("article_ref")))

    logger.info(f"[DocAI Ingest] {req.blob_name} → {count} chunk, {pages} pagine, {arts} art.refs")
    return {
        "blob_name":      req.blob_name,
        "chunks_indexed": count,
        "pages_covered":  pages,
        "article_refs":   arts,
        "cached":         False,
        "ocr_used":       True,
    }


@app.get("/api/documents/docai-chunks")
async def get_docai_chunks(
    blob_name: str = Query(..., description="Nome del blob nel bucket"),
    db: Session = Depends(get_db),
):
    """
    Ritorna i chunk salvati con coordinate spaziali — senza vettori embedding.
    Usato dal frontend per il sistema di evidenziazione geometrica (pixel-perfect).

    Risposta:
      {
        "indexed": true,
        "total": 312,
        "chunks": [
          { "id": 0, "text": "...", "page": 5,
            "coords": {"x0":72,"y0":120,"x1":524,"y1":145,"page_width":595,"page_height":842},
            "article_ref": "Articolo 30" },
          ...
        ]
      }
    """
    rows = load_embeddings(db, blob_name)
    if not rows:
        return {"indexed": False, "total": 0, "chunks": []}

    chunks = [
        {
            "id":          r.get("chunk_index", i),
            "text":        r.get("text", ""),
            "page":        r.get("page_number"),
            "coords":      r.get("coordinates"),    # {x0,y0,x1,y1,page_width,page_height}
            "article_ref": r.get("article_ref"),
        }
        for i, r in enumerate(rows)
    ]
    return {"indexed": True, "total": len(chunks), "chunks": chunks}


# ── Gemini Proxy (frontend can't call Gemini REST directly — key restricted to server) ──

class GeminiProxyRequest(BaseModel):
    # Simple format: { prompt, max_tokens, temperature }
    prompt: Optional[str] = None
    model: Optional[str] = None
    max_tokens: int = 2048
    temperature: float = 0.3
    # Native format: { system_instruction, contents, generationConfig }
    system_instruction: Optional[dict] = None
    contents: Optional[list] = None
    generationConfig: Optional[dict] = None

@app.post("/api/gemini")
async def gemini_proxy(req: GeminiProxyRequest):
    """
    Proxy frontend Gemini calls through the backend.
    Supports two call formats:
      1. Simple: { prompt, max_tokens, temperature }
      2. Native: { system_instruction, contents, generationConfig }
    """
    if not _VERTEX_AI_AVAILABLE:
        raise HTTPException(status_code=503, detail="google-genai not installed")
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=503, detail="GEMINI_API_KEY not configured")
    if not req.prompt and not req.contents:
        raise HTTPException(status_code=422, detail="Either 'prompt' or 'contents' is required")

    primary_model = req.model or VERTEX_MODEL
    client = _genai.Client(api_key=GEMINI_API_KEY)

    # Build config and contents once, reuse across model attempts
    if req.contents:
        native_contents = req.contents
        gen_cfg = req.generationConfig or {}
        if req.system_instruction:
            sys_parts = req.system_instruction.get('parts', [])
            sys_text = ' '.join(p.get('text', '') for p in sys_parts)
            config = _genai_types.GenerateContentConfig(
                max_output_tokens=gen_cfg.get('maxOutputTokens', 6144),
                temperature=gen_cfg.get('temperature', 0.1),
                system_instruction=sys_text,
            )
        else:
            config = _genai_types.GenerateContentConfig(
                max_output_tokens=gen_cfg.get('maxOutputTokens', 6144),
                temperature=gen_cfg.get('temperature', 0.1),
            )
        def _make_call(m):
            return lambda: client.models.generate_content(
                model=m, contents=native_contents, config=config)
    else:
        config = _genai_types.GenerateContentConfig(
            max_output_tokens=req.max_tokens,
            temperature=req.temperature,
        )
        def _make_call(m):
            return lambda: client.models.generate_content(
                model=m, contents=req.prompt, config=config)

    # Try primary model, then fallbacks on model-not-found errors
    models_to_try = [primary_model] + [
        m for m in _PROXY_FALLBACK_MODELS if m != primary_model
    ]
    last_err: Exception = Exception("No models available")
    for model_id in models_to_try:
        try:
            response = await asyncio.to_thread(_make_call(model_id))
            if model_id != primary_model:
                logger.info(f"[GeminiProxy] fallback OK → {model_id}")
            return {"text": response.text or ""}
        except Exception as e:
            err_str = str(e).lower()
            is_model_err = any(k in err_str for k in [
                "not found", "404", "invalid", "does not exist",
                "unknown model", "model_not_found", "not supported"
            ])
            if is_model_err:
                logger.warning(f"[GeminiProxy] model '{model_id}' unavailable, trying next…")
                last_err = e
                continue
            # Non-model error (auth, quota, network…) — fail fast
            logger.error(f"[GeminiProxy] error with model '{model_id}': {e}")
            raise HTTPException(status_code=500, detail=str(e))

    logger.error(f"[GeminiProxy] all models exhausted. Last error: {last_err}")
    raise HTTPException(status_code=503, detail=f"Nessun modello disponibile: {last_err}")


# ═══════════════════════════════════════════════════════════════════
# UPLOAD — Fase 3/4/5
# ═══════════════════════════════════════════════════════════════════

@app.post("/api/upload/request-url")
async def request_upload_url(req: UploadRequestBody):
    """
    Fase 3 — Step 1: genera una Signed URL PUT per l'upload diretto
    browser → GCS (il file non transita dal server).

    Struttura cartella GCS: uploads/{user_id}/{timestamp}_{filename}
    Restituisce: { signed_url, blob_name, expires_at }
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    # Validazione dimensione file (max 50 MB)
    MAX_SIZE = 50 * 1024 * 1024
    if req.file_size and req.file_size > MAX_SIZE:
        raise HTTPException(status_code=413, detail="File troppo grande (max 50 MB)")

    # Tipi MIME consentiti
    ALLOWED_MIME = {
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/plain",
    }
    if req.content_type not in ALLOWED_MIME:
        raise HTTPException(status_code=415, detail=f"Tipo file non consentito: {req.content_type}")

    ts         = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name  = _sanitize_filename(req.filename)
    blob_name  = f"uploads/{req.user_id}/{ts}_{safe_name}"
    expires_at = (datetime.datetime.utcnow() + datetime.timedelta(minutes=req.expiration_minutes)).isoformat() + "Z"

    try:
        signed_url = gcs.generate_upload_signed_url(
            blob_name=blob_name,
            content_type=req.content_type,
            expiration_minutes=req.expiration_minutes,
        )
        logger.info(f"[Upload] URL generata per {blob_name} (user={req.user_id})")
        return {
            "signed_url": signed_url,
            "blob_name":  blob_name,
            "expires_at": expires_at,
        }
    except Exception as e:
        logger.error(f"[Upload] Errore generazione URL: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload/confirm")
async def confirm_upload(req: UploadConfirmBody, db: Session = Depends(get_db)):
    """
    Fase 5 — Dopo che il browser ha completato il PUT su GCS,
    chiama questo endpoint per salvare il record nel DB.
    Restituisce il record completo con ID e stato 'in_attesa'.
    """
    try:
        record = create_upload(
            db=db,
            user_id=req.user_id,
            blob_name=req.blob_name,
            original_filename=req.original_filename,
            file_size=req.file_size,
            mime_type=req.mime_type,
        )
        logger.info(f"[Upload] Confermato id={record.id} blob={req.blob_name}")
        return record.to_dict()
    except Exception as e:
        logger.error(f"[Upload] Errore conferma: {e}")
        raise HTTPException(status_code=500, detail=str(e))


LOCAL_UPLOADS_DIR = os.path.join(os.path.dirname(__file__), "local_uploads")
os.makedirs(LOCAL_UPLOADS_DIR, exist_ok=True)

@app.post("/api/upload/direct")
async def upload_direct(
    file: UploadFile = File(...),
    user_id: str = Form("utente_anonimo"),
    db: Session = Depends(get_db),
):
    """
    Upload diretto: file salvato localmente sul server.
    Non richiede permessi GCS di scrittura.
    """
    ALLOWED_MIME = {
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/plain",
    }
    MAX_SIZE = 50 * 1024 * 1024  # 50 MB

    content_type = file.content_type or "application/octet-stream"
    data = await file.read()
    if len(data) > MAX_SIZE:
        raise HTTPException(status_code=413, detail="File troppo grande (max 50 MB)")

    ts        = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = _sanitize_filename(file.filename or "document")
    local_filename = f"{ts}_{safe_name}"
    local_path     = os.path.join(LOCAL_UPLOADS_DIR, local_filename)
    blob_name      = f"local/{local_filename}"   # virtual blob_name for DB

    with open(local_path, "wb") as f_out:
        f_out.write(data)
    logger.info(f"[Upload Direct] Salvato localmente: {local_path} ({len(data)} bytes)")

    record = create_upload(
        db=db,
        user_id=user_id,
        blob_name=blob_name,
        original_filename=file.filename or safe_name,
        file_size=len(data),
        mime_type=content_type,
    )
    logger.info(f"[Upload Direct] DB id={record.id}")
    result = record.to_dict()
    result["local_path"] = local_path
    return result


@app.get("/api/upload/local/{filename}")
async def serve_local_upload(filename: str):
    """Serve a locally-stored uploaded file."""
    safe = _sanitize_filename(filename)
    path = os.path.join(LOCAL_UPLOADS_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File non trovato")
    import mimetypes
    mime, _ = mimetypes.guess_type(path)
    return FileResponse(path, media_type=mime or "application/octet-stream",
                        headers={"Content-Disposition": f'inline; filename="{safe}"'})


@app.get("/api/upload/local/{filename}/extract-text")
async def extract_local_text(filename: str):
    """Extract text from a locally-stored Word/Excel/TXT file for display in viewer."""
    safe = _sanitize_filename(filename)
    path = os.path.join(LOCAL_UPLOADS_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File non trovato")

    ext = safe.rsplit(".", 1)[-1].lower() if "." in safe else ""
    try:
        if ext in ("docx",):
            from docx import Document as DocxDocument
            doc = DocxDocument(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n\n".join(paragraphs)
            return {"filename": safe, "type": "word", "text": text, "paragraphs": len(paragraphs)}

        elif ext in ("doc",):
            # .doc (old binary format) — extract via antiword if available, else fallback
            try:
                import subprocess
                result = subprocess.run(["antiword", path], capture_output=True, text=True, timeout=10)
                text = result.stdout if result.returncode == 0 else ""
            except Exception:
                text = ""
            if not text:
                text = "[Formato .doc non supportato per preview — scarica il file per aprirlo in Word]"
            return {"filename": safe, "type": "word_legacy", "text": text}

        elif ext in ("xlsx",):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"=== Foglio: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append("\t".join(cells))
            wb.close()
            text = "\n".join(lines)
            return {"filename": safe, "type": "excel", "text": text}

        elif ext in ("xls",):
            text = "[Formato .xls non supportato per preview — scarica il file per aprirlo in Excel]"
            return {"filename": safe, "type": "excel_legacy", "text": text}

        elif ext in ("txt", "csv", "md"):
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
            return {"filename": safe, "type": "text", "text": text}

        else:
            return {"filename": safe, "type": "unknown", "text": f"[Anteprima non disponibile per .{ext}]"}

    except Exception as e:
        logger.error(f"[ExtractText] {safe}: {e}")
        raise HTTPException(status_code=500, detail=f"Errore estrazione testo: {e}")


@app.get("/api/upload/files")
async def get_uploads(
    user_id: Optional[str] = Query(None,  description="Filtra per utente"),
    status:  Optional[str] = Query(None,  description="Filtra per stato"),
    limit:   int            = Query(200,   description="Max risultati"),
    db: Session = Depends(get_db),
):
    """
    Fase 4A — Lista file caricati (con filtri opzionali).
    Admin: nessun user_id → tutti i file.
    Utente: user_id → solo i propri.
    """
    records = list_uploads(db=db, user_id=user_id, status=status, limit=limit)
    return [r.to_dict() for r in records]


@app.patch("/api/upload/files/{upload_id}/status")
async def patch_upload_status(
    upload_id: int,
    req: StatusUpdateBody,
    db: Session = Depends(get_db),
):
    """
    Fase 4A (Admin) — Aggiorna lo stato di un file:
    in_attesa | approvato | rifiutato
    """
    VALID = {"in_attesa", "approvato", "rifiutato"}
    if req.status not in VALID:
        raise HTTPException(status_code=422, detail=f"Stato non valido. Valori: {VALID}")

    record = update_status(db=db, upload_id=upload_id, status=req.status, notes=req.notes)
    if not record:
        raise HTTPException(status_code=404, detail=f"Upload id={upload_id} non trovato")
    logger.info(f"[Upload] Status id={upload_id} → {req.status}")
    return record.to_dict()


@app.delete("/api/upload/files/{upload_id}")
async def delete_upload_record(
    upload_id: int,
    remove_from_gcs: bool = Query(False, description="Se True elimina anche il file da GCS"),
    db: Session = Depends(get_db),
):
    """
    Fase 4A (Admin) — Elimina il record dal DB.
    Con remove_from_gcs=true elimina anche il file fisico da GCS.
    """
    from database import Upload as UploadModel
    record = db.query(UploadModel).filter(UploadModel.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail=f"Upload id={upload_id} non trovato")

    blob_name = record.blob_name
    deleted = delete_upload(db=db, upload_id=upload_id)

    file_removed = False
    if blob_name.startswith("local/"):
        # File stored locally — delete from disk
        local_filename = blob_name[len("local/"):]
        local_path = os.path.join(LOCAL_UPLOADS_DIR, local_filename)
        if os.path.exists(local_path):
            try:
                os.remove(local_path)
                file_removed = True
                logger.info(f"[Upload] File locale eliminato: {local_path}")
            except Exception as e:
                logger.warning(f"[Upload] Impossibile eliminare file locale: {e}")
    elif remove_from_gcs and gcs:
        try:
            file_removed = gcs.delete_blob(blob_name)
            logger.info(f"[Upload] Blob GCS eliminato: {blob_name}")
        except Exception as e:
            logger.warning(f"[Upload] Impossibile eliminare da GCS: {e}")

    return {"deleted": deleted, "blob_name": blob_name, "file_removed": file_removed}


@app.get("/api/upload/files/{upload_id}/preview-url")
async def get_preview_url(
    upload_id: int,
    expiration: int = Query(30, description="Durata URL in minuti"),
    db: Session = Depends(get_db),
):
    """
    Fase 4B — Genera una Signed URL GET per visualizzare il file
    (iframe PDF o Google Docs Viewer per Word/Excel).
    Restituisce anche viewer_url pronto per <iframe src=...>
    """
    if not gcs:
        raise HTTPException(status_code=503, detail="GCS client non inizializzato")

    from database import Upload as UploadModel
    record = db.query(UploadModel).filter(UploadModel.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail=f"Upload id={upload_id} non trovato")

    try:
        signed_url = gcs.generate_signed_url(record.blob_name, expiration_minutes=expiration)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Determina il viewer appropriato
    mime = (record.mime_type or "").lower()
    if "pdf" in mime:
        viewer_url = signed_url          # iframe diretto
        viewer_type = "pdf"
    else:
        # Google Docs Viewer per Word/Excel (nessun plugin necessario)
        import urllib.parse
        encoded = urllib.parse.quote(signed_url, safe="")
        viewer_url  = f"https://docs.google.com/viewer?url={encoded}&embedded=true"
        viewer_type = "google_docs"

    return {
        "id":           upload_id,
        "blob_name":    record.blob_name,
        "signed_url":   signed_url,
        "viewer_url":   viewer_url,
        "viewer_type":  viewer_type,
        "expires_in_minutes": expiration,
    }


# ── Serve Frontend (optional) ──────────────────────────────────────

# Unico file sorgente — nella stessa cartella del backend
FRONTEND_HTML = os.path.join(
    os.path.dirname(__file__), 'cipp_e_IAPP_UmbertoMottolaSaas_v5_50_fix.html'
)

@app.get("/")
async def root():
    """Serve the CIPP/E SaaS frontend HTML."""
    if os.path.exists(FRONTEND_HTML):
        # Read and return directly — avoids Starlette FileResponse sendfile() issues on macOS
        with open(FRONTEND_HTML, "rb") as f:
            content = f.read()
        return Response(content=content, media_type="text/html; charset=utf-8")
    return {
        "app": "CIPP/E Legal SaaS API",
        "version": "1.0.0",
        "docs": "/docs",
        "endpoints": [
            "GET  /api/health",
            "GET  /api/documents",
            "GET  /api/documents/signed-url?blob_name=...",
            "POST /api/documents/extract-text",
            "POST /api/documents/highlights",
            "GET  /api/documents/metadata?blob_name=...",
            "GET  /api/documents/ai-context?blob_name=...",
        ]
    }


# ── Run ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=os.getenv("HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", "8000")),
        reload=True
    )
